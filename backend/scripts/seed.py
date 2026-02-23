import os
import re
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import select

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.append(str(BASE_DIR))

load_dotenv(BASE_DIR / ".env")

from app.db.session import SessionLocal
from app.core.security import hash_password, verify_password
from app.models.security import User, UserRole
from app.models.core import (
    Manufacturer,
    EquipmentType,
    EquipmentCategory,
    Warehouse,
    Personnel,
    Location,
    MeasurementUnit,
    MainEquipment,
)
from app.models.assemblies import Assembly
from app.models.operations import AssemblyItem

MAIN_EQUIPMENT_CODE_RE = re.compile(r"^\d+(?:\.\d+)*$")


def find_main_equipment_xlsx() -> Path | None:
    linux_path = Path("/mnt/data/mill_equipment.xlsx")
    if linux_path.exists():
        return linux_path

    for path in Path("C:/Users").glob("*/Downloads/mill_equipment.xlsx"):
        if path.exists():
            return path

    return None


def seed_main_equipment_from_excel(db) -> dict[str, int | str | None]:
    xlsx_path = find_main_equipment_xlsx()
    if not xlsx_path:
        return {
            "created": 0,
            "updated": 0,
            "restored": 0,
            "skipped_invalid": 0,
            "skipped_missing_parent": 0,
            "path": None,
        }

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb["Оборудование фабрики"] if "Оборудование фабрики" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 3:
        return {
            "created": 0,
            "updated": 0,
            "restored": 0,
            "skipped_invalid": 0,
            "skipped_missing_parent": 0,
            "path": str(xlsx_path),
        }

    headers = ["" if cell is None else str(cell).strip() for cell in rows[1]]
    code_idx = 0
    name_idx = 1
    subtype_idx = 2
    automation_idx = 3

    if len(headers) > 0 and headers[0].strip().lower() == "код":
        code_idx = 0

    candidates_by_code: dict[str, dict] = {}
    skipped_invalid = 0
    pending_group_label: str | None = None

    for row in rows[2:]:
        values = list(row)
        while len(values) <= max(code_idx, name_idx, subtype_idx, automation_idx):
            values.append(None)

        code_raw = values[code_idx]
        code = "" if code_raw is None else str(code_raw).strip()
        if not code:
            skipped_invalid += 1
            continue

        if not MAIN_EQUIPMENT_CODE_RE.match(code):
            # Excel contains helper group rows like "— По конструкции:"
            # without a numeric code; use them to name missing parent groups.
            cleaned = code.strip("—- ").rstrip(":").strip()
            if cleaned:
                pending_group_label = cleaned
            skipped_invalid += 1
            continue

        level = code.count(".") + 1
        primary_name = "" if values[name_idx] is None else str(values[name_idx]).strip()
        subtype_name = "" if values[subtype_idx] is None else str(values[subtype_idx]).strip()
        name_raw = primary_name or subtype_name

        name = "" if name_raw is None else str(name_raw).strip()
        if not name:
            skipped_invalid += 1
            continue

        automation_raw = values[automation_idx]
        automation = "" if automation_raw is None else str(automation_raw).strip()
        meta_data = {"automation_params": automation} if automation else None

        parent_code = code.rsplit(".", 1)[0] if "." in code else None
        candidates_by_code[code] = {
            "code": code,
            "name": name,
            "level": level,
            "parent_code": parent_code,
            "meta_data": meta_data,
            "is_placeholder": False,
        }

        if pending_group_label and parent_code:
            parent_level = parent_code.count(".") + 1
            existing_parent = candidates_by_code.get(parent_code)
            if existing_parent is None:
                candidates_by_code[parent_code] = {
                    "code": parent_code,
                    "name": pending_group_label,
                    "level": parent_level,
                    "parent_code": parent_code.rsplit(".", 1)[0] if "." in parent_code else None,
                    "meta_data": None,
                    "is_placeholder": True,
                }
            elif existing_parent.get("is_placeholder") and existing_parent["name"].startswith("Группа "):
                existing_parent["name"] = pending_group_label

    # Ensure intermediate parent prefixes exist, so deep nodes are not dropped.
    for code in list(candidates_by_code.keys()):
        parts = code.split(".")
        if len(parts) <= 1:
            continue
        for idx in range(1, len(parts)):
            prefix = ".".join(parts[:idx])
            if prefix not in candidates_by_code:
                parent_code = ".".join(parts[: idx - 1]) if idx > 1 else None
                candidates_by_code[prefix] = {
                    "code": prefix,
                    "name": f"Группа {prefix}",
                    "level": idx,
                    "parent_code": parent_code,
                    "meta_data": None,
                    "is_placeholder": True,
                }

    candidates = sorted(
        candidates_by_code.values(),
        key=lambda item: (item["level"], tuple(int(part) for part in item["code"].split("."))),
    )

    code_to_id: dict[str, int] = {
        item.code: item.id for item in db.scalars(select(MainEquipment)).all() if item.code
    }

    created = 0
    updated = 0
    restored = 0
    skipped_missing_parent = 0

    for candidate in candidates:
        parent_id = None
        if candidate["parent_code"] is not None:
            parent_id = code_to_id.get(candidate["parent_code"])
            if parent_id is None:
                parent = db.scalar(
                    select(MainEquipment).where(
                        MainEquipment.code == candidate["parent_code"],
                        MainEquipment.is_deleted == False,
                    )
                )
                if parent:
                    parent_id = parent.id
                    code_to_id[candidate["parent_code"]] = parent.id
            if parent_id is None:
                skipped_missing_parent += 1
                continue

        active = db.scalar(
            select(MainEquipment).where(
                MainEquipment.code == candidate["code"],
                MainEquipment.is_deleted == False,
            )
        )
        deleted = None
        if active is None:
            deleted = db.scalar(
                select(MainEquipment).where(
                    MainEquipment.code == candidate["code"],
                    MainEquipment.is_deleted == True,
                )
            )

        if active is not None:
            changed = False
            if active.name != candidate["name"]:
                active.name = candidate["name"]
                changed = True
            if active.level != candidate["level"]:
                active.level = candidate["level"]
                changed = True
            if active.parent_id != parent_id:
                active.parent_id = parent_id
                changed = True
            if active.meta_data != candidate["meta_data"]:
                active.meta_data = candidate["meta_data"]
                changed = True
            if changed:
                updated += 1
            code_to_id[candidate["code"]] = active.id
            continue

        if deleted is not None:
            deleted.is_deleted = False
            deleted.deleted_at = None
            deleted.deleted_by_id = None
            deleted.name = candidate["name"]
            deleted.level = candidate["level"]
            deleted.parent_id = parent_id
            deleted.meta_data = candidate["meta_data"]
            restored += 1
            code_to_id[candidate["code"]] = deleted.id
            continue

        item = MainEquipment(
            name=candidate["name"],
            parent_id=parent_id,
            level=candidate["level"],
            code=candidate["code"],
            meta_data=candidate["meta_data"],
        )
        db.add(item)
        db.flush()
        created += 1
        code_to_id[candidate["code"]] = item.id

    return {
        "created": created,
        "updated": updated,
        "restored": restored,
        "skipped_invalid": skipped_invalid,
        "skipped_missing_parent": skipped_missing_parent,
        "path": str(xlsx_path),
    }


def run():
    db = SessionLocal()
    try:
        admin_username = os.getenv("SEED_ADMIN_USERNAME", "admin")
        admin_password = os.getenv("SEED_ADMIN_PASSWORD", "admin12345")
        admin = db.scalar(select(User).where(User.username == admin_username))
        if not admin:
            admin = User(
                username=admin_username,
                password_hash=hash_password(admin_password),
                role=UserRole.admin,
            )
            db.add(admin)
        else:
            if admin.is_deleted:
                admin.is_deleted = False
                admin.deleted_at = None
                admin.deleted_by_id = None
            if admin.role != UserRole.admin:
                admin.role = UserRole.admin
            if not verify_password(admin_password, admin.password_hash):
                admin.password_hash = hash_password(admin_password)

        siemens = db.scalar(
            select(Manufacturer).where(Manufacturer.name == "Siemens", Manufacturer.is_deleted == False)
        )
        if not siemens:
            siemens = Manufacturer(name="Siemens", country="Germany")
            db.add(siemens)

        wh = db.scalar(
            select(Warehouse).where(Warehouse.name == "Склад 1", Warehouse.is_deleted == False)
        )
        if not wh:
            wh = Warehouse(name="Склад 1")
            db.add(wh)

        location = db.scalar(
            select(Location).where(Location.name == "Локация 1", Location.is_deleted == False)
        )
        if not location:
            location = Location(name="Локация 1")
            db.add(location)

        assembly = db.scalar(
            select(Assembly).where(Assembly.name == "Сборка 1", Assembly.is_deleted == False)
        )
        if not assembly:
            assembly = Assembly(name="Сборка 1", location=location)
            db.add(assembly)
            db.flush()

        et = db.scalar(
            select(EquipmentType).where(
                EquipmentType.nomenclature_number == "PLC-001", EquipmentType.is_deleted == False
            )
        )
        if not et:
            et = EquipmentType(
                name="PLC базовый",
                nomenclature_number="PLC-001",
                manufacturer=siemens,
                is_channel_forming=True,
                channel_count=16,
                ai_count=16,
                di_count=0,
                ao_count=0,
                do_count=0,
                is_network=False,
                meta_data={"unit_price_rub": 100000},
            )
            db.add(et)
            db.flush()

        if et and assembly:
            assembly_item = db.scalar(
                select(AssemblyItem).where(
                    AssemblyItem.assembly_id == assembly.id,
                    AssemblyItem.equipment_type_id == et.id,
                )
            )
            if not assembly_item:
                assembly_item = AssemblyItem(
                    assembly_id=assembly.id,
                    equipment_type_id=et.id,
                    quantity=2,
                )
                db.add(assembly_item)

        category_names = [
            "ПЛК",
            "Реле",
            "HMI",
            "Блок питания",
            "Преобразователь",
            "Программируемое реле",
        ]
        for name in category_names:
            category = db.scalar(
                select(EquipmentCategory).where(
                    EquipmentCategory.name == name
                )
            )
            if not category:
                category = EquipmentCategory(name=name)
                db.add(category)
            elif category.is_deleted:
                category.is_deleted = False
                category.deleted_at = None
                category.deleted_by_id = None

        measurement_units_tree = {
            "Температура": ["Кельвин (K)", "градус Цельсия (°C)", "градус Фаренгейта (°F)", "градус Ренкина (°R)"],
            "Уровень": ["Метр (м)", "миллиметр (мм)", "процент (%)", "дюйм (in)", "фут (ft)"],
            "Расход": [
                "м3/ч",
                "л/мин",
                "л/с",
                "кг/ч",
                "т/сут",
                "Нм3/ч (нормальный)",
                "См3/ч (стандартный)",
                "SCFM",
                "GPM",
            ],
            "Давление": [
                "Паскаль (Па)",
                "кПа",
                "МПа",
                "бар",
                "кгс/см2",
                "мм рт. ст.",
                "мм вод. ст.",
                "psi (psia, psig, psid)",
                "Торр",
                "мбар",
            ],
            "Плотность": [
                "кг/м3",
                "г/см3",
                "градус API",
                "градус Брикса (°Bx)",
                "градус Плато (°P)",
                "градус Боме (°Be)",
                "градус Эксле (°Oe)",
                "градус Баллинга",
            ],
            "pH/ОВП/Оксиметрия": ["ед. pH", "милливольт (мВ)", "мг/л", "мкг/л", "% насыщения"],
            "Электропараметры": [
                "Ампер (А)",
                "мА",
                "Вольт (В)",
                "мВ",
                "Ватт (Вт)",
                "кВтч",
                "мкСм/см",
                "мСм/см",
                "Омсм",
                "МОм*см",
                "ppm (TDS)",
            ],
            "Вибрация": ["мкм", "mils", "мм/с (RMS/Peak)", "ips", "м/с2", "g", "дБ", "Гц"],
            "Концентрация газов": ["% Vol", "ppm", "ppb", "% LEL (НКПР)", "мг/м3", "мкг/м3"],
            "Энкодеры": ["имп/об (PPR)", "отсчетов (CPR)", "бит (bit)", "мкм", "LPI", "DPI", "угловые секунды", "градусы"],
            "Свет": ["Кандела (кд)", "Люмен (лм)", "Люкс (лк)", "foot-candle (fc)", "кд/м2 (нит)", "мкВт/см2"],
            "Влажность": ["% RH (относительная)", "°C Td (точка росы)", "г/м3", "г/кг", "ppmV"],
            "Масса и вес": ["кг", "т", "фунт (lb)", "Ньютон (Н)", "кН", "мВ/В (РКП)", "деления (e, d)"],
        }
        created = 0
        skipped = 0

        def get_or_create_unit(name: str, parent_id: int | None, sort_order: int | None) -> MeasurementUnit:
            unit = db.scalar(
                select(MeasurementUnit).where(
                    MeasurementUnit.name == name,
                    MeasurementUnit.parent_id == parent_id,
                    MeasurementUnit.is_deleted == False,
                )
            )
            if unit:
                return unit
            unit = db.scalar(
                select(MeasurementUnit).where(
                    MeasurementUnit.name == name,
                    MeasurementUnit.parent_id == parent_id,
                    MeasurementUnit.is_deleted == True,
                )
            )
            if unit:
                unit.is_deleted = False
                unit.deleted_at = None
                unit.deleted_by_id = None
                unit.sort_order = sort_order
                return unit
            new_unit = MeasurementUnit(name=name, parent_id=parent_id, sort_order=sort_order)
            db.add(new_unit)
            return new_unit

        for root_index, (root_name, units) in enumerate(measurement_units_tree.items(), start=1):
            root_unit = get_or_create_unit(root_name, None, root_index)
            if root_unit.id is None:
                created += 1
            else:
                skipped += 1
            db.flush()
            for child_index, child_name in enumerate(units, start=1):
                child_unit = get_or_create_unit(child_name, root_unit.id, child_index)
                if child_unit.id is None:
                    created += 1
                else:
                    skipped += 1

        main_equipment_stats = seed_main_equipment_from_excel(db)

        db.commit()
        existing_personnel = db.scalar(select(Personnel).limit(1))
        if not existing_personnel:
            person = Personnel(
                first_name="Иван",
                last_name="Иванов",
                position="Инженер",
                department="АСУ",
                organisation="EQM Demo"
            )
            db.add(person)
            db.commit()
        print(f"Measurement units seed: created={created}, skipped={skipped}")
        print(
            "Main equipment seed: "
            f"created={main_equipment_stats['created']}, "
            f"updated={main_equipment_stats['updated']}, "
            f"restored={main_equipment_stats['restored']}, "
            f"skipped_invalid={main_equipment_stats['skipped_invalid']}, "
            f"skipped_missing_parent={main_equipment_stats['skipped_missing_parent']}, "
            f"source={main_equipment_stats['path']}"
        )
        print("Seed completed.")
    finally:
        db.close()


if __name__ == "__main__":
    run()

