from __future__ import annotations

import csv
import json
from datetime import date, datetime
from io import BytesIO, StringIO, TextIOWrapper
from typing import Any, Iterable, Sequence

from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def detect_upload_format(file: UploadFile) -> str:
    content_type = (file.content_type or "").lower()
    filename = (file.filename or "").lower()
    if "spreadsheetml" in content_type or filename.endswith(".xlsx"):
        return "xlsx"
    if content_type in {"text/csv", "application/csv", "text/plain"} or filename.endswith(".csv"):
        return "csv"
    raise HTTPException(status_code=400, detail="Unsupported file format")


def read_tabular_rows(file: UploadFile, file_format: str | None = None) -> tuple[list[str], list[list[Any]]]:
    effective_format = file_format or detect_upload_format(file)
    if effective_format == "xlsx":
        workbook = load_workbook(file.file, data_only=True)
        sheet = workbook["DATA"] if "DATA" in workbook.sheetnames else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    else:
        text_stream = TextIOWrapper(file.file, encoding="utf-8-sig")
        reader = csv.reader(text_stream)
        rows = list(reader)
    if not rows:
        return [], []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return headers, [list(row) if row is not None else [] for row in rows[1:]]


def build_workbook_buffer(
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    readme_lines: Sequence[str] | None = None,
) -> BytesIO:
    workbook = Workbook()
    if readme_lines:
        workbook.active.title = "README"
        readme = workbook.active
        for line in readme_lines:
            readme.append([line])
        sheet = workbook.create_sheet("DATA")
    else:
        sheet = workbook.active
        sheet.title = "DATA"
    sheet.append(list(headers))
    for row in rows:
        sheet.append([serialize_cell(value) for value in row])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_csv_buffer(headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> StringIO:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(list(headers))
    for row in rows:
        writer.writerow([serialize_cell(value) for value in row])
    buffer.seek(0)
    return buffer


def build_export_response(
    *,
    filename_prefix: str,
    file_format: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    readme_lines: Sequence[str] | None = None,
) -> StreamingResponse:
    if file_format == "csv":
        buffer = build_csv_buffer(headers, rows)
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_prefix}.csv"'},
        )

    buffer = build_workbook_buffer(headers, rows, readme_lines=readme_lines)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_prefix}.xlsx"'},
    )


def build_template_response(
    *,
    filename_prefix: str,
    headers: Sequence[str],
    readme_lines: Sequence[str],
) -> StreamingResponse:
    buffer = build_workbook_buffer(headers, [], readme_lines=readme_lines)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_prefix}.xlsx"'},
    )


def serialize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def is_blank_row(row: Sequence[Any]) -> bool:
    return not any(value not in (None, "") for value in row)


def row_to_mapping(headers: Sequence[str], row: Sequence[Any]) -> dict[str, Any]:
    mapping: dict[str, Any] = {}
    for index, header in enumerate(headers):
        mapping[normalize_header(header)] = row[index] if index < len(row) else None
    return mapping


def as_optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def as_required_str(value: Any, *, field: str) -> str:
    text = as_optional_str(value)
    if not text:
        raise ValueError(f"{field} is required")
    return text


def as_optional_int(value: Any) -> int | None:
    text = as_optional_str(value)
    if text is None:
        return None
    return int(float(text))


def as_optional_float(value: Any) -> float | None:
    text = as_optional_str(value)
    if text is None:
        return None
    return float(text.replace(",", "."))


def as_optional_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y", "да"}:
        return True
    if normalized in {"false", "0", "no", "n", "нет"}:
        return False
    raise ValueError("Invalid boolean value")


def as_optional_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    return date.fromisoformat(str(value).strip())
