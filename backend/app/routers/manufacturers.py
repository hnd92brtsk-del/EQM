import csv
from datetime import datetime
from io import BytesIO, StringIO, TextIOWrapper

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.core.dependencies import get_db, require_read_access, require_write_access
from app.core.pagination import paginate
from app.core.query import apply_search, apply_sort, apply_date_filters
from app.core.audit import add_audit_log, model_to_dict
from app.models.core import Manufacturer
from app.models.security import User
from app.schemas.common import Pagination
from app.schemas.manufacturers import (
    ImportIssue,
    ImportReport,
    ManufacturerOut,
    ManufacturerCreate,
    ManufacturerUpdate,
)

router = APIRouter()


def _normalize_name(value: str) -> str:
    return " ".join(value.split()).strip().casefold()


def _build_template_xlsx() -> BytesIO:
    workbook = Workbook()
    workbook.active.title = "README"
    readme = workbook.active
    readme.append(["Manufacturers import template"])
    readme.append(["Columns: name (required), country (required)"])
    readme.append(["Duplicates are skipped by normalized name."])
    readme.append(["Normalization: trim + collapse spaces + casefold"])
    readme.append(["Import is create-only."])

    data = workbook.create_sheet("DATA")
    data.append(["name", "country"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_export_xlsx(items: list[Manufacturer]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DATA"
    sheet.append(["name", "country"])
    for item in items:
        sheet.append([item.name, item.country])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_export_csv(items: list[Manufacturer]) -> StringIO:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["name", "country"])
    for item in items:
        writer.writerow([item.name, item.country])
    buffer.seek(0)
    return buffer


def _read_xlsx_rows(file: UploadFile) -> tuple[list[str], list[list[str | None]]]:
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook["DATA"] if "DATA" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def _read_csv_rows(file: UploadFile) -> tuple[list[str], list[list[str | None]]]:
    text_stream = TextIOWrapper(file.file, encoding="utf-8-sig")
    reader = csv.reader(text_stream)
    rows = list(reader)
    if not rows:
        return [], []
    headers = [value.strip() for value in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def _detect_format(file: UploadFile) -> str:
    content_type = (file.content_type or "").lower()
    filename = (file.filename or "").lower()
    if "spreadsheetml" in content_type or filename.endswith(".xlsx"):
        return "xlsx"
    if content_type in {"text/csv", "application/csv", "text/plain"} or filename.endswith(".csv"):
        return "csv"
    raise HTTPException(status_code=400, detail="Unsupported file format")


@router.get("/", response_model=Pagination[ManufacturerOut])
def list_manufacturers(
    page: int = 1,
    page_size: int = 50,
    q: str | None = None,
    sort: str | None = None,
    is_deleted: bool | None = None,
    include_deleted: bool = False,
    created_at_from: datetime | None = None,
    created_at_to: datetime | None = None,
    updated_at_from: datetime | None = None,
    updated_at_to: datetime | None = None,
    db=Depends(get_db),
    user: User = Depends(require_read_access()),
):
    query = select(Manufacturer)
    if is_deleted is None:
        if not include_deleted:
            query = query.where(Manufacturer.is_deleted == False)
    else:
        query = query.where(Manufacturer.is_deleted == is_deleted)

    query = apply_search(query, q, [Manufacturer.name, Manufacturer.country])
    query = apply_date_filters(query, Manufacturer, created_at_from, created_at_to, updated_at_from, updated_at_to)
    query = apply_sort(query, Manufacturer, sort)

    total, items = paginate(query, db, page, page_size)
    return Pagination(items=items, page=page, page_size=page_size, total=total)


@router.get("/export")
def export_manufacturers(
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
    is_deleted: bool | None = None,
    include_deleted: bool = False,
    db=Depends(get_db),
    user: User = Depends(require_read_access()),
):
    query = select(Manufacturer)
    if is_deleted is None:
        if not include_deleted:
            query = query.where(Manufacturer.is_deleted == False)
    else:
        query = query.where(Manufacturer.is_deleted == is_deleted)
    items = db.scalars(query).all()

    if format == "csv":
        buffer = _build_export_csv(items)
        headers = {
            "Content-Disposition": 'attachment; filename="manufacturers.csv"'
        }
        return StreamingResponse(buffer, media_type="text/csv", headers=headers)

    buffer = _build_export_xlsx(items)
    headers = {
        "Content-Disposition": 'attachment; filename="manufacturers.xlsx"'
    }
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/template")
def download_template(
    format: str = Query(default="xlsx", pattern="^(xlsx)$"),
    user: User = Depends(require_read_access()),
):
    buffer = _build_template_xlsx()
    headers = {
        "Content-Disposition": 'attachment; filename="manufacturers-template.xlsx"'
    }
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import", response_model=ImportReport)
def import_manufacturers(
    file: UploadFile = File(...),
    format: str | None = Query(default=None, pattern="^(csv|xlsx)$"),
    dry_run: bool = True,
    db=Depends(get_db),
    current_user: User = Depends(require_write_access()),
):
    effective_format = format or _detect_format(file)
    if effective_format == "csv":
        headers, data_rows = _read_csv_rows(file)
    else:
        headers, data_rows = _read_xlsx_rows(file)

    if not headers:
        raise HTTPException(status_code=400, detail="Empty file")

    header_map = {header.strip().lower(): idx for idx, header in enumerate(headers)}
    if "name" not in header_map:
        raise HTTPException(status_code=400, detail="Missing 'name' column")
    if "country" not in header_map:
        raise HTTPException(status_code=400, detail="Missing 'country' column")

    existing_names = db.scalars(
        select(Manufacturer.name).where(Manufacturer.is_deleted == False)
    ).all()
    seen = {_normalize_name(name) for name in existing_names}

    report = ImportReport(
        total_rows=0,
        created=0,
        skipped_duplicates=0,
        errors=[],
        warnings=[],
    )

    to_create: list[Manufacturer] = []
    for index, row in enumerate(data_rows, start=2):
        if row is None:
            continue
        row_values = list(row)
        if not any(value not in (None, "") for value in row_values):
            continue
        report.total_rows += 1

        raw_name = row_values[header_map["name"]] if header_map["name"] < len(row_values) else None
        raw_country = (
            row_values[header_map["country"]] if header_map["country"] < len(row_values) else None
        )
        name = str(raw_name).strip() if raw_name is not None else ""
        country = str(raw_country).strip() if raw_country is not None else ""

        if not name:
            report.errors.append(
                ImportIssue(row=index, field="name", message="Name is required")
            )
            continue
        if not country:
            report.errors.append(
                ImportIssue(row=index, field="country", message="Country is required")
            )
            continue

        normalized = _normalize_name(name)
        if normalized in seen:
            report.skipped_duplicates += 1
            report.warnings.append(
                ImportIssue(row=index, field="name", message="Duplicate name skipped")
            )
            continue

        seen.add(normalized)
        manufacturer = Manufacturer(name=name, country=country)
        to_create.append(manufacturer)

    report.created = len(to_create)
    if not dry_run and to_create:
        db.add_all(to_create)
        db.flush()
        for manufacturer in to_create:
            add_audit_log(
                db,
                actor_id=current_user.id,
                action="CREATE",
                entity="manufacturers",
                entity_id=manufacturer.id,
                before=None,
                after=model_to_dict(manufacturer),
            )
        db.commit()

    return report


@router.get("/{manufacturer_id}", response_model=ManufacturerOut)
def get_manufacturer(
    manufacturer_id: int,
    include_deleted: bool = False,
    db=Depends(get_db),
    user: User = Depends(require_read_access()),
):
    query = select(Manufacturer).where(Manufacturer.id == manufacturer_id)
    if not include_deleted:
        query = query.where(Manufacturer.is_deleted == False)
    manufacturer = db.scalar(query)
    if not manufacturer:
        raise HTTPException(status_code=404, detail="Manufacturer not found")
    return manufacturer


@router.post("/", response_model=ManufacturerOut)
def create_manufacturer(
    payload: ManufacturerCreate,
    db=Depends(get_db),
    current_user: User = Depends(require_write_access()),
):
    existing = db.scalar(
        select(Manufacturer).where(
            Manufacturer.name == payload.name, Manufacturer.is_deleted == False
        )
    )
    if existing:
        raise HTTPException(status_code=400, detail="Manufacturer already exists")

    manufacturer = Manufacturer(name=payload.name, country=payload.country)
    db.add(manufacturer)
    db.flush()

    add_audit_log(
        db,
        actor_id=current_user.id,
        action="CREATE",
        entity="manufacturers",
        entity_id=manufacturer.id,
        before=None,
        after=model_to_dict(manufacturer),
    )

    db.commit()
    db.refresh(manufacturer)
    return manufacturer


@router.patch("/{manufacturer_id}", response_model=ManufacturerOut)
def update_manufacturer(
    manufacturer_id: int,
    payload: ManufacturerUpdate,
    db=Depends(get_db),
    current_user: User = Depends(require_write_access()),
):
    manufacturer = db.scalar(select(Manufacturer).where(Manufacturer.id == manufacturer_id))
    if not manufacturer:
        raise HTTPException(status_code=404, detail="Manufacturer not found")

    before = model_to_dict(manufacturer)

    if payload.name is not None:
        manufacturer.name = payload.name
    if payload.country is not None:
        manufacturer.country = payload.country

    add_audit_log(
        db,
        actor_id=current_user.id,
        action="UPDATE",
        entity="manufacturers",
        entity_id=manufacturer.id,
        before=before,
        after=model_to_dict(manufacturer),
    )

    db.commit()
    db.refresh(manufacturer)
    return manufacturer


@router.put("/{manufacturer_id}", response_model=ManufacturerOut)
def update_manufacturer_legacy(
    manufacturer_id: int,
    payload: ManufacturerUpdate,
    db=Depends(get_db),
    current_user: User = Depends(require_write_access()),
):
    return update_manufacturer(manufacturer_id, payload, db, current_user)


@router.delete("/{manufacturer_id}")
def delete_manufacturer(
    manufacturer_id: int,
    db=Depends(get_db),
    current_user: User = Depends(require_write_access()),
):
    manufacturer = db.scalar(select(Manufacturer).where(Manufacturer.id == manufacturer_id))
    if not manufacturer:
        raise HTTPException(status_code=404, detail="Manufacturer not found")

    before = model_to_dict(manufacturer)
    manufacturer.is_deleted = True
    manufacturer.deleted_at = datetime.utcnow()
    manufacturer.deleted_by_id = current_user.id

    add_audit_log(
        db,
        actor_id=current_user.id,
        action="DELETE",
        entity="manufacturers",
        entity_id=manufacturer.id,
        before=before,
        after=model_to_dict(manufacturer),
    )

    db.commit()
    return {"status": "ok"}


@router.post("/{manufacturer_id}/restore", response_model=ManufacturerOut)
def restore_manufacturer(
    manufacturer_id: int,
    db=Depends(get_db),
    current_user: User = Depends(require_write_access()),
):
    manufacturer = db.scalar(select(Manufacturer).where(Manufacturer.id == manufacturer_id))
    if not manufacturer:
        raise HTTPException(status_code=404, detail="Manufacturer not found")

    before = model_to_dict(manufacturer)
    manufacturer.is_deleted = False
    manufacturer.deleted_at = None
    manufacturer.deleted_by_id = None

    add_audit_log(
        db,
        actor_id=current_user.id,
        action="RESTORE",
        entity="manufacturers",
        entity_id=manufacturer.id,
        before=before,
        after=model_to_dict(manufacturer),
    )

    db.commit()
    db.refresh(manufacturer)
    return manufacturer
